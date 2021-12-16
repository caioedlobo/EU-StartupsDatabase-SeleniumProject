import os

import openpyxl
import pytest

# -*- coding: utf-8 -*-

@pytest.mark.usefixtures("setup")
class BaseClass:
    path = "StartupsDatabase.xlsx"

    def getBaseURL(self):
        return self.driver.get("https://www.eu-startups.com/directory/")

    def setExcel(self):
        os.chdir("ExcelSheet")

        book = openpyxl.load_workbook(self.path)
        sheet = book.active

        sheet.cell(row=1, column=1).value = "Company Name"
        sheet.cell(row=1, column=2).value = "Category"
        sheet.cell(row=1, column=3).value = "Based In"
        sheet.cell(row=1, column=4).value = "Tags"
        sheet.cell(row=1, column=5).value = "Founded"
        sheet.cell(row=1, column=6).value = "Description"

    def setExcelSheet(self, initial, startup_list):
        book = openpyxl.load_workbook(self.path)
        sheet = book.active


        sheet.cell(row=initial, column=1).value = startup_list[0]
        sheet.cell(row=initial, column=2).value = startup_list[1]
        sheet.cell(row=initial, column=3).value = startup_list[2]
        sheet.cell(row=initial, column=4).value = startup_list[3]
        sheet.cell(row=initial, column=5).value = startup_list[4]
        sheet.cell(row=initial, column=6).value = startup_list[5]

        book.save(self.path)



